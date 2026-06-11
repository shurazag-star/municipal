from decimal import Decimal

from openpyxl import Workbook

from municipal_agent.budget_sources import BudgetSource
from municipal_agent.excel_parser import parse_xlsx_finance_report
from municipal_agent.row_classification import ExcelRowType


def test_excel_parser_finds_result_sheet_totals_duplicates_and_residual(tmp_path):
    path = tmp_path / "finance.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.append(
        [
            "Наименование",
            "Код целевой программы",
            "Код мероприятия",
            "Код объекта",
            "Наименование объекта по адресному перечню",
            "2026 всего",
            "2026 бюджет субъекта РФ",
            "2026 местный бюджет",
            "2027 всего",
            "2027 бюджет субъекта РФ",
            "2027 местный бюджет",
            "2028 всего",
            "2028 бюджет субъекта РФ",
            "2028 местный бюджет",
        ]
    )
    ws.append(["Муниципальная программа Развитие ЖКХ", "01", "", "", "", 2253220255.91, 0, 0, 1776791196.12, 0, 0, 780689180, 0, 0])
    ws.append(["", "", "01.01", "CHERUSTI", "ВЗУ Черусти", 90555380, 78330390, 12224990, 0, 0, 0, 0, 0, 0])
    ws.append(["", "", "02.03", "0000000000.0000000000", "", 1000, 0, 1000, 0, 0, 0, 0, 0, 0])
    ws.append(["Итого:", "", "", "", "", 2253220255.91, 0, 0, 1776791196.12, 0, 0, 780689180, 0, 0])
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.sheet_name == "Результат"
    assert parsed.program_totals[2026] == Decimal("2253220255.91")
    assert sum(1 for row in parsed.rows if row.row_type == ExcelRowType.FINAL_TOTAL_ROW) == 1
    assert parsed.total_without_double_count()[2026] == Decimal("2253220255.91")
    assert parsed.object_groups[0].total_by_year()[2026] == Decimal("90555380.00")
    assert parsed.object_groups[1].status == "UNASSIGNED_RESIDUAL"
    assert parsed.object_groups[1].funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("1000.00")


def test_excel_parser_supports_budget_roster_plan_columns_and_type_codes(tmp_path):
    path = tmp_path / "budget_roster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.merge_cells("B8:M8")
    ws.merge_cells("N8:O8")
    ws.merge_cells("P8:Q8")
    ws.merge_cells("R8:S8")
    ws.merge_cells("T8:U8")
    ws.merge_cells("V8:X8")
    ws.merge_cells("Y8:Z8")
    ws.merge_cells("AA8:AB8")
    ws["B8"] = "Наименование"
    ws["N8"] = "ЦСР"
    ws["P8"] = "Тип средств"
    ws["R8"] = "Мероприятие"
    ws["T8"] = "Объект"
    ws["V8"] = "План на 2026 год"
    ws["Y8"] = "План на 2027 год"
    ws["AA8"] = "План на 2028год"
    ws["B9"] = 1
    ws["N9"] = 2
    ws["P9"] = 3
    ws["R9"] = 4
    ws["T9"] = 5
    ws["V9"] = 6
    ws["Y9"] = 7
    ws["AA9"] = 8

    ws["C10"] = 'Муниципальная программа "Развитие инженерной инфраструктуры"'
    ws["N10"] = "1000000000"
    ws["V10"] = 181000
    ws["Y10"] = 304000
    ws["AA10"] = 500000

    ws["F20"] = "Реконструкция ВЗУ Птицефабрика"
    ws["N20"] = "10102S4090"
    ws["P20"] = "900100"
    ws["R20"] = "101020100000000"
    ws["T20"] = "1000010939.5327942989"
    ws["V20"] = 17160000
    ws["Y20"] = 65780000
    ws["AA20"] = 0

    ws["F21"] = "Реконструкция ВЗУ Птицефабрика"
    ws["N21"] = "10102S4090"
    ws["P21"] = "900302"
    ws["R21"] = "101020100000000"
    ws["T21"] = "1000010939.5327942989"
    ws["V21"] = 300000
    ws["Y21"] = 400000
    ws["AA21"] = 500000

    ws["B30"] = "Итого:"
    ws["V30"] = 181000
    ws["Y30"] = 304000
    ws["AA30"] = 500000
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.program_totals[2026] == Decimal("181000.00")
    assert parsed.final_totals[2028] == Decimal("500000.00")
    assert parsed.rows[0].funding == {}
    group = parsed.object_groups[0]
    assert group.rows[0].object_name == "Реконструкция ВЗУ Птицефабрика"
    assert group.funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("17160000.00")
    assert group.funding[(2026, BudgetSource.REGIONAL_BUDGET)] == Decimal("300000.00")
    assert group.funding[(2028, BudgetSource.REGIONAL_BUDGET)] == Decimal("500000.00")


def test_excel_parser_supports_relative_budget_roster_plan_years(tmp_path):
    path = tmp_path / "relative_budget_roster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws["B2"] = "с 03.12.2025 по 07.05.2026"
    ws.merge_cells("B8:M8")
    ws.merge_cells("N8:O8")
    ws.merge_cells("P8:Q8")
    ws.merge_cells("R8:S8")
    ws.merge_cells("T8:U8")
    ws.merge_cells("V8:X8")
    ws.merge_cells("Y8:Z8")
    ws.merge_cells("AA8:AB8")
    ws["B8"] = "Наименование"
    ws["N8"] = "ЦСР"
    ws["P8"] = "Тип средств"
    ws["R8"] = "Мероприятие"
    ws["T8"] = "Объект"
    ws["V8"] = "План на 1 год"
    ws["Y8"] = "План на 2 год"
    ws["AA8"] = "План на 3 год"
    ws["B9"] = 1
    ws["N9"] = 2
    ws["P9"] = 3
    ws["R9"] = 4
    ws["T9"] = 5
    ws["V9"] = 6
    ws["Y9"] = 7
    ws["AA9"] = 8

    ws["C10"] = 'Муниципальная программа "Развитие инженерной инфраструктуры"'
    ws["N10"] = "1000000000"
    ws["V10"] = 181000
    ws["Y10"] = 304000
    ws["AA10"] = 500000

    ws["F20"] = "Реконструкция ВЗУ Птицефабрика"
    ws["N20"] = "10102S4090"
    ws["P20"] = "900100"
    ws["R20"] = "101020100000000"
    ws["T20"] = "1000010939.5327942989"
    ws["V20"] = 17160000
    ws["Y20"] = 65780000
    ws["AA20"] = 0

    ws["F21"] = "Реконструкция ВЗУ Птицефабрика"
    ws["N21"] = "10102S4090"
    ws["P21"] = "900302"
    ws["R21"] = "101020100000000"
    ws["T21"] = "1000010939.5327942989"
    ws["V21"] = 300000
    ws["Y21"] = 400000
    ws["AA21"] = 500000
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.program_totals[2026] == Decimal("181000.00")
    group = parsed.object_groups[0]
    assert group.funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("17160000.00")
    assert group.funding[(2027, BudgetSource.LOCAL_BUDGET)] == Decimal("65780000.00")
    assert group.funding[(2028, BudgetSource.REGIONAL_BUDGET)] == Decimal("500000.00")


def test_excel_parser_supports_current_financial_year_plan_columns(tmp_path):
    path = tmp_path / "Отчет о финанс. мероп. цел. прогр.№10 на 09.06.2026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws["A1"] = "Отчет о финансировании мероприятий целевых программ"
    ws["A2"] = "по 09 июня 2026 г."
    ws["A6"] = "Наименование"
    ws["B6"] = "Классификация"
    ws["F6"] = "Мероприятие"
    ws["G6"] = "Наименование Объекта по АП"
    ws["I6"] = "Тип средств"
    ws["K6"] = "Утверждено плановые назначения"
    ws["Q6"] = "Поставлено на учет БО (заключено контрактов, договоров)"
    ws["R6"] = "Фактически исполнено"
    ws["U6"] = "Свободных ЛБО"
    ws["B7"] = "Код цел. программы._x000D_\nКод мероприятия"
    ws["K7"] = "Всего на текущий финансовый год"
    ws["R7"] = "Всего"
    ws["N8"] = "средства бюджета субъекта"
    ws["O8"] = "средства местного бюджета"
    ws["S8"] = "средства бюджета субъекта"
    ws["T8"] = "средства местного бюджета"
    for idx, col in enumerate(["A", "B", "C", "D", "E", "F", "G", "I", "K", "N", "O", "Q", "R", "S", "T", "U"], 1):
        ws[f"{col}9"] = idx

    ws["A10"] = '10 - Муниципальная программа "Развитие инженерной инфраструктуры"'
    ws["K10"] = 2_155_969_664.33
    ws["N10"] = 1_760_494_100
    ws["O10"] = 438_124_644.33

    ws["A18"] = "Строительство ВЗУ со станцией водоочистки"
    ws["B18"] = "10102S4090"
    ws["F18"] = "101020100000000"
    ws["G18"] = "Строительство ВЗУ со станцией водоочистки и разводящими сетями водоснабжения г.о. Шатура, р.п. Черусти"
    ws["I18"] = "900100"
    ws["K18"] = 90_555_380
    ws["N18"] = 78_330_390
    ws["O18"] = 12_224_990
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.target_years == [2026]
    assert parsed.program_totals[2026] == Decimal("2155969664.33")
    group = parsed.object_groups[0]
    assert group.parent_activity_code == "101020100000000"
    assert group.funding[(2026, BudgetSource.REGIONAL_BUDGET)] == Decimal("78330390.00")
    assert group.funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("12224990.00")
    assert group.total_by_year()[2026] == Decimal("90555380.00")


def test_excel_parser_keeps_separate_year_header_row_before_numbering(tmp_path):
    path = tmp_path / "separate_year_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.merge_cells("B8:M8")
    ws.merge_cells("N8:O8")
    ws.merge_cells("P8:Q8")
    ws.merge_cells("R8:S8")
    ws.merge_cells("T8:U8")
    ws.merge_cells("V9:X9")
    ws.merge_cells("Y9:AA9")
    ws.merge_cells("AB9:AD9")
    ws["B8"] = "Наименование"
    ws["N8"] = "ЦСР"
    ws["P8"] = "Тип средств"
    ws["R8"] = "Мероприятие"
    ws["T8"] = "Объект"
    ws["V9"] = 2026
    ws["Y9"] = 2027
    ws["AB9"] = 2028
    for start in ("V", "Y", "AB"):
        ws[f"{start}10"] = "Всего"
    ws["W10"] = "Средства бюджета субъекта РФ"
    ws["X10"] = "Средства бюджета муниципального округа"
    ws["Z10"] = "Средства бюджета субъекта РФ"
    ws["AA10"] = "Средства бюджета муниципального округа"
    ws["AC10"] = "Средства бюджета субъекта РФ"
    ws["AD10"] = "Средства бюджета муниципального округа"
    ws["B11"] = 1
    ws["N11"] = 2
    ws["P11"] = 3
    ws["R11"] = 4
    ws["T11"] = 5
    ws["V11"] = 6
    ws["W11"] = 7
    ws["X11"] = 8
    ws["Y11"] = 9
    ws["Z11"] = 10
    ws["AA11"] = 11
    ws["AB11"] = 12
    ws["AC11"] = 13
    ws["AD11"] = 14

    ws["C12"] = 'Муниципальная программа "Развитие инженерной инфраструктуры"'
    ws["N12"] = "1000000000"
    ws["V12"] = 181000
    ws["Y12"] = 304000
    ws["AB12"] = 500000

    ws["F20"] = "Реконструкция ВЗУ Птицефабрика"
    ws["N20"] = "10102S4090"
    ws["P20"] = "900100"
    ws["R20"] = "101020100000000"
    ws["T20"] = "1000010939.5327942989"
    ws["V20"] = 17160000
    ws["W20"] = 0
    ws["X20"] = 17160000
    ws["Y20"] = 65780000
    ws["Z20"] = 0
    ws["AA20"] = 65780000
    ws["AB20"] = 500000
    ws["AC20"] = 500000
    ws["AD20"] = 0
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.target_years == [2026, 2027, 2028]
    assert parsed.program_totals[2026] == Decimal("181000.00")
    group = parsed.object_groups[0]
    assert group.funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("17160000.00")
    assert group.funding[(2027, BudgetSource.LOCAL_BUDGET)] == Decimal("65780000.00")
    assert group.funding[(2028, BudgetSource.REGIONAL_BUDGET)] == Decimal("500000.00")


def test_excel_parser_supports_plain_year_headers_and_activity_totals(tmp_path):
    path = tmp_path / "activity_budget_roster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.merge_cells("B8:M8")
    ws.merge_cells("N8:O8")
    ws.merge_cells("P8:Q8")
    ws.merge_cells("R8:S8")
    ws["B8"] = "Наименование"
    ws["N8"] = "ЦСР"
    ws["P8"] = "Тип средств"
    ws["R8"] = "Мероприятие"
    ws["T8"] = "КОСГУ"
    ws["U8"] = "СубКОСГУ"
    ws["V8"] = "2026 год"
    ws["Y8"] = "2027 год"
    ws["AA8"] = "2028 год"
    ws["B9"] = 1
    ws["N9"] = 2
    ws["P9"] = 3
    ws["R9"] = 4
    ws["T9"] = 5
    ws["U9"] = 6
    ws["V9"] = 7
    ws["Y9"] = 8
    ws["AA9"] = 9

    ws["C10"] = 'Муниципальная программа "Развитие институтов гражданского общества"'
    ws["N10"] = "1300000000"
    ws["V10"] = 184730729.5
    ws["Y10"] = 126648829.5
    ws["AA10"] = 126681369.5

    ws["F20"] = "Информирование населения в печатных СМИ"
    ws["N20"] = "1310100821"
    ws["P20"] = "900100"
    ws["R20"] = "131010500000000"
    ws["V20"] = 2_000_000
    ws["Y20"] = 2_000_000
    ws["AA20"] = 2_000_000

    ws["F21"] = "Прочие работы, услуги"
    ws["N21"] = "1310100821"
    ws["P21"] = "900100"
    ws["R21"] = "131010500000000"
    ws["T21"] = "226"
    ws["V21"] = 2_000_000
    ws["Y21"] = 2_000_000
    ws["AA21"] = 2_000_000

    ws["F22"] = "Информирование населения в печатных СМИ"
    ws["N22"] = "1310100821"
    ws["P22"] = "900302"
    ws["R22"] = "131010500000000"
    ws["V22"] = 300_000
    ws["Y22"] = 400_000
    ws["AA22"] = 500_000
    wb.save(path)

    parsed = parse_xlsx_finance_report(path)

    assert parsed.program_totals[2026] == Decimal("184730729.50")
    assert parsed.program_totals[2028] == Decimal("126681369.50")
    assert parsed.target_years == [2026, 2027, 2028]
    activity_rows = [row for row in parsed.rows if row.row_type == ExcelRowType.ACTIVITY_AGGREGATE_ROW]
    assert len(activity_rows) == 3
    assert len(parsed.object_groups) == 1
    group = parsed.object_groups[0]
    assert group.status == "ACTIVITY_AGGREGATE"
    assert group.object_code == "131010500000000"
    assert group.object_name == "Информирование населения в печатных СМИ"
    assert group.funding[(2026, BudgetSource.LOCAL_BUDGET)] == Decimal("2000000.00")
    assert group.funding[(2027, BudgetSource.LOCAL_BUDGET)] == Decimal("2000000.00")
    assert group.funding[(2028, BudgetSource.REGIONAL_BUDGET)] == Decimal("500000.00")
    assert group.total_by_year()[2026] == Decimal("2300000.00")
