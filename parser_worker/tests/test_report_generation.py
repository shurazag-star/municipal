from pathlib import Path

import pytest
from openpyxl import load_workbook

from municipal_agent.reports import generate_reconciliation_artifacts


ROOT = Path(__file__).resolve().parents[2]
SAMPLES = ROOT / "sample_documents"
DOCX = SAMPLES / "проект изменений МП_март_2026 (10).docx"
XLSX = SAMPLES / "Отчет_о_финансировании_мероприятий_целевых_программ+_Расширенный.xlsx"
PDF = SAMPLES / "2. № 2291 от 16.10.2025.pdf"
pytestmark = pytest.mark.skipif(
    not DOCX.exists() or not XLSX.exists() or not PDF.exists(),
    reason="real sample documents are not committed to the deploy repository",
)


def test_generates_real_document_reconciliation_artifacts(tmp_path):
    artifacts = generate_reconciliation_artifacts(
        docx_path=DOCX,
        xlsx_path=XLSX,
        pdf_path=PDF,
        output_dir=tmp_path,
    )

    mapping = artifacts["mapping_report_json"]
    html = artifacts["control_sums_report_html"]
    xlsx = artifacts["change_report_xlsx"]

    assert mapping.exists()
    assert html.exists()
    assert xlsx.exists()
    assert "PROGRAM_TOTAL_DIFF" in html.read_text(encoding="utf-8")
    assert "ВЗУ Черусти" in html.read_text(encoding="utf-8")

    workbook = load_workbook(xlsx)
    sheet = workbook["Расхождения"]
    assert sheet.max_row == 4
    assert sheet["A2"].value == 2026
    assert sheet["E2"].value == -42881704.09
