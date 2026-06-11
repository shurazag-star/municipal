from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
import re
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Tuple

from openpyxl import load_workbook

from .budget_sources import BudgetSource, normalize_budget_source
from .money import parse_money_to_rub, quantize_rub
from .normalization import normalize_name
from .row_classification import ExcelRowType, classify_excel_row


FundingKey = Tuple[int, BudgetSource]
AmountColumn = Tuple[int, Optional[BudgetSource], bool]


@dataclass
class ExcelFinanceRow:
    row_number: int
    row_type: ExcelRowType
    parent_activity_code: str = ""
    object_code: str = ""
    object_name: str = ""
    funding: Dict[FundingKey, Decimal] = field(default_factory=dict)
    raw_values: Dict[str, object] = field(default_factory=dict)
    explicit_zero_target: bool = False


@dataclass
class ExcelObjectGroup:
    group_key: str
    status: str
    parent_activity_code: str = ""
    object_code: str = ""
    object_name: str = ""
    rows: List[ExcelFinanceRow] = field(default_factory=list)
    funding: Dict[FundingKey, Decimal] = field(default_factory=dict)
    explicit_zero_target: bool = False

    def total_by_year(self) -> Dict[int, Decimal]:
        totals: Dict[int, Decimal] = {}
        for (year, _source), amount in self.funding.items():
            totals[year] = quantize_rub(totals.get(year, Decimal("0")) + amount)
        return totals


@dataclass
class ParsedExcelReport:
    sheet_name: str
    rows: List[ExcelFinanceRow]
    program_totals: Dict[int, Decimal]
    final_totals: Dict[int, Decimal]
    object_groups: List[ExcelObjectGroup]
    target_years: List[int] = field(default_factory=list)

    def total_without_double_count(self) -> Dict[int, Decimal]:
        if self.program_totals:
            return dict(self.program_totals)
        totals: Dict[int, Decimal] = {}
        for group in self.object_groups:
            for year, amount in group.total_by_year().items():
                totals[year] = quantize_rub(totals.get(year, Decimal("0")) + amount)
        return totals


def group_excel_object_rows(rows: Iterable[ExcelFinanceRow]) -> List[ExcelObjectGroup]:
    groups: Dict[str, ExcelObjectGroup] = {}
    for row in rows:
        identity = _group_identity(row)
        if identity is None:
            continue

        group_key, status, object_code, object_name = identity
        group = groups.setdefault(
            group_key,
            ExcelObjectGroup(
                group_key=group_key,
                status=status,
                parent_activity_code=row.parent_activity_code.strip(),
                object_code=object_code,
                object_name=object_name,
            ),
        )
        group.rows.append(row)
        group.explicit_zero_target = group.explicit_zero_target or row.explicit_zero_target
        for key, amount in row.funding.items():
            group.funding[key] = quantize_rub(group.funding.get(key, Decimal("0")) + amount)

    return list(groups.values())


def _group_identity(row: ExcelFinanceRow) -> Optional[Tuple[str, str, str, str]]:
    if row.row_type == ExcelRowType.UNASSIGNED_RESIDUAL_ROW:
        group_key = f"UNASSIGNED_RESIDUAL::{row.parent_activity_code}::{row.row_number}"
        return group_key, "UNASSIGNED_RESIDUAL", row.object_code.strip(), row.object_name.strip()

    if row.row_type == ExcelRowType.OBJECT_LEAF_ROW:
        if _service_object_row(row):
            return None
        object_name = row.object_name.strip()
        object_code = row.object_code.strip()
        object_identity = object_code or normalize_name(object_name)
        group_key = "::".join([row.parent_activity_code.strip(), object_identity, normalize_name(object_name)])
        return group_key, "GROUPED_OBJECT", object_code, object_name

    if row.row_type == ExcelRowType.ACTIVITY_AGGREGATE_ROW and _activity_total_row(row):
        object_name = row.object_name.strip() or str(row.raw_values.get("Наименование") or "").strip()
        object_code = row.parent_activity_code.strip()
        object_identity = object_code or normalize_name(object_name)
        group_key = "::".join([row.parent_activity_code.strip(), object_identity, normalize_name(object_name)])
        return group_key, "ACTIVITY_AGGREGATE", object_code, object_name

    return None


def _activity_total_row(row: ExcelFinanceRow) -> bool:
    return bool(row.parent_activity_code.strip()) and bool(row.funding) and not _kosgu_value_present(row.raw_values)


def _service_object_row(row: ExcelFinanceRow) -> bool:
    return normalize_name(row.object_name) in {"рейтинг"}


def _kosgu_value_present(values: Mapping[str, object]) -> bool:
    for header, value in values.items():
        header_norm = header.lower().replace("ё", "е")
        if "косгу" in header_norm and "суб" not in header_norm and not _blank(value):
            return True
    return False


def parse_xlsx_finance_report(path: str | Path) -> ParsedExcelReport:
    workbook = load_workbook(filename=path, data_only=True)
    sheet_name = "Результат" if "Результат" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    data_start_row, headers = _detect_header(worksheet)
    base_year = _infer_relative_plan_base_year(worksheet) or _infer_year_from_path(path)
    amount_columns = _detect_amount_columns(headers, base_year=base_year)
    rows: List[ExcelFinanceRow] = []
    program_totals: Dict[int, Decimal] = {}
    final_totals: Dict[int, Decimal] = {}

    for row_cells in worksheet.iter_rows(min_row=data_start_row, values_only=False):
        values = _row_values(row_cells, headers)
        normalized = _normalize_row_values(values)
        row_type = classify_excel_row(normalized)
        funding = _extract_funding(row_cells, amount_columns, row_source=_row_budget_source(values))
        finance_row = ExcelFinanceRow(
            row_number=row_cells[0].row,
            row_type=row_type,
            parent_activity_code=normalized["measure_code"],
            object_code=normalized["object_code"],
            object_name=normalized["object_name"] or (normalized["name"] if row_type == ExcelRowType.ACTIVITY_AGGREGATE_ROW else ""),
            funding=funding,
            raw_values=values,
            explicit_zero_target=row_type == ExcelRowType.OBJECT_LEAF_ROW and bool(amount_columns) and not funding,
        )
        rows.append(finance_row)
        if row_type == ExcelRowType.PROGRAM_TOTAL_ROW:
            program_totals.update(_extract_total_columns(row_cells, amount_columns))
        elif row_type == ExcelRowType.FINAL_TOTAL_ROW:
            final_totals.update(_extract_total_columns(row_cells, amount_columns))

    return ParsedExcelReport(
        sheet_name=sheet_name,
        rows=rows,
        program_totals=program_totals,
        final_totals=final_totals,
        object_groups=group_excel_object_rows(rows),
        target_years=sorted({year for year, _source, _row_source_required in amount_columns.values()}),
    )


def _detect_header(worksheet) -> Tuple[int, Dict[int, str]]:
    for row_cells in worksheet.iter_rows(values_only=False):
        texts = ["" if cell.value is None else str(cell.value).strip() for cell in row_cells]
        if any("наименование" in text.lower() for text in texts):
            header_start = row_cells[0].row
            header_rows = [row_cells]
            data_start = header_start + 1
            for candidate in worksheet.iter_rows(min_row=header_start + 1, max_row=header_start + 4, values_only=False):
                candidate_texts = ["" if cell.value is None else str(cell.value).strip() for cell in candidate]
                filled = [text for text in candidate_texts if text]
                if _looks_like_year_header_row(candidate_texts):
                    header_rows.append(candidate)
                    data_start = candidate[0].row + 1
                    continue
                if filled and all(_is_header_number(text) for text in filled):
                    data_start = candidate[0].row + 1
                    break
                joined = " ".join(candidate_texts).lower().replace("ё", "е")
                if any(marker in joined for marker in ["код ", "в том числе", "средства", "всего на"]):
                    header_rows.append(candidate)
                    data_start = candidate[0].row + 1
                else:
                    data_start = candidate[0].row
                    break
            headers: Dict[int, str] = {}
            max_columns = max(len(row) for row in header_rows)
            for idx in range(max_columns):
                parts = []
                for header_row in header_rows:
                    if idx < len(header_row):
                        value = _merged_cell_value(header_row[idx])
                        if value not in (None, ""):
                            parts.append(str(value).replace("_x000D_", " ").replace("\n", " ").strip())
                if parts:
                    headers[idx] = " ".join(parts)
            return data_start, headers
    raise ValueError("Excel header row was not found")


def _is_header_number(text: str) -> bool:
    return str(text).strip().isdigit()


def _looks_like_year_header_row(texts: Iterable[str]) -> bool:
    return any(_looks_like_year_header_cell(text) for text in texts)


def _looks_like_year_header_cell(text: str) -> bool:
    normalized = str(text).strip().lower().replace("ё", "е")
    normalized = re.sub(r"\s+", " ", normalized)
    if re.fullmatch(r"20\d{2}(?:\s*(?:г|год|года))?", normalized):
        return True
    return bool(re.search(r"\bплан\s+на\s+(?:20\d{2}|[123]\s*год)", normalized))


def _merged_cell_value(cell):
    if cell.value not in (None, ""):
        return cell.value
    worksheet = cell.parent
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def _row_values(row_cells, headers: Mapping[int, str]) -> Dict[str, object]:
    values: Dict[str, object] = {}
    for idx, header in headers.items():
        value = row_cells[idx].value
        if header not in values or _blank(values[header]):
            values[header] = value
    return values


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_row_values(values: Mapping[str, object]) -> Dict[str, str]:
    name = _lookup_name(values)
    object_code = _lookup_object_code(values)
    object_name = _lookup_object_name(values)
    if not object_name and object_code:
        object_name = name
    return {
        "name": name,
        "program_code": _lookup_program_code(values),
        "measure_code": _lookup_measure_code(values),
        "object_code": object_code,
        "object_name": object_name,
    }


def _lookup_by_predicate(values: Mapping[str, object], predicate) -> str:
    for header, value in values.items():
        header_norm = header.lower().replace("ё", "е")
        if predicate(header_norm):
            return "" if value is None else str(value).strip()
    return ""


def _lookup_name(values: Mapping[str, object]) -> str:
    return _lookup_by_predicate(values, lambda header: header.startswith("наименование") and "объект" not in header)


def _lookup_program_code(values: Mapping[str, object]) -> str:
    return _lookup_by_predicate(values, lambda header: "код цел" in header or "код программы" in header)


def _lookup_measure_code(values: Mapping[str, object]) -> str:
    value = _lookup_by_predicate(values, lambda header: header.strip().startswith("мероприятие"))
    if value:
        return value
    return _lookup_by_predicate(values, lambda header: "код мероприятия" in header and "код цел" not in header)


def _lookup_object_code(values: Mapping[str, object]) -> str:
    return _lookup_by_predicate(values, lambda header: "объект" in header and "наименование" not in header)


def _lookup_object_name(values: Mapping[str, object]) -> str:
    return _lookup_by_predicate(values, lambda header: "наименование" in header and "объект" in header)


def _lookup_budget_source(values: Mapping[str, object]) -> str:
    return _lookup_by_predicate(values, lambda header: "тип средств" in header or "источник" in header)


def _row_budget_source(values: Mapping[str, object]) -> BudgetSource:
    source = normalize_budget_source(_lookup_budget_source(values))
    return source


def _detect_amount_columns(headers: Mapping[int, str], base_year: Optional[int] = None) -> Dict[int, AmountColumn]:
    columns: Dict[int, AmountColumn] = {}
    current_year: Optional[int] = None
    in_plan_columns = True
    for idx, header in sorted(headers.items()):
        text = header.lower().replace("ё", "е")
        if any(marker in text for marker in ["поставлено", "фактически", "% исполнения"]):
            in_plan_columns = False
        if not in_plan_columns:
            continue
        match = re.search(r"(20\d{2})", text)
        if match:
            current_year = int(match.group(1))
        current_year_match = _looks_like_current_financial_year_header(text) and base_year is not None
        if not match and current_year_match:
            current_year = base_year
        relative_match = re.search(r"план\s+на\s+([123])\s*год", text)
        if not match and relative_match and base_year is not None:
            current_year = base_year + int(relative_match.group(1)) - 1
        source = normalize_budget_source(text)
        year_header = match or relative_match or current_year_match
        if year_header and "всего" in text:
            columns[idx] = (current_year, None, False)
        elif year_header and _looks_like_plan_total_header(text):
            columns[idx] = (current_year, None, True)
        elif match and _looks_like_plain_year_total_header(text):
            columns[idx] = (current_year, None, False)
        elif current_year is not None and source is not BudgetSource.UNKNOWN:
            columns[idx] = (current_year, source, False)
    return columns


def _infer_relative_plan_base_year(worksheet) -> Optional[int]:
    years: List[int] = []
    max_row = min(worksheet.max_row, 10)
    max_col = min(worksheet.max_column, 10)
    for row_cells in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        for value in row_cells:
            years.extend(int(match) for match in re.findall(r"20\d{2}", "" if value is None else str(value)))
    return max(years) if years else None


def _infer_year_from_path(path: str | Path) -> Optional[int]:
    years = [int(match) for match in re.findall(r"20\d{2}", str(path))]
    return max(years) if years else None


def _looks_like_current_financial_year_header(text: str) -> bool:
    return "текущий финансовый год" in text


def _looks_like_plan_total_header(text: str) -> bool:
    return "план на" in text or re.search(r"\bплан\b", text)


def _looks_like_plain_year_total_header(text: str) -> bool:
    normalized = re.sub(r"[^0-9a-zа-я]+", " ", text.lower().replace("ё", "е")).strip()
    return bool(re.fullmatch(r"20\d{2}(?: год(?:а)?)?", normalized))


def _extract_funding(
    row_cells,
    amount_columns: Mapping[int, AmountColumn],
    row_source: BudgetSource = BudgetSource.UNKNOWN,
) -> Dict[FundingKey, Decimal]:
    specific: Dict[FundingKey, Decimal] = {}
    totals: Dict[int, Decimal] = {}
    for idx, (year, source, row_source_required) in amount_columns.items():
        amount = parse_money_to_rub(row_cells[idx].value)
        if amount == Decimal("0.00"):
            continue
        if source is None and row_source is BudgetSource.UNKNOWN:
            if row_source_required:
                continue
            totals[year] = amount
        else:
            source = row_source if source is None else source
            specific[(year, source)] = amount
    if specific:
        return specific
    return {(year, BudgetSource.UNKNOWN): amount for year, amount in totals.items()}


def _extract_total_columns(row_cells, amount_columns: Mapping[int, AmountColumn]) -> Dict[int, Decimal]:
    totals: Dict[int, Decimal] = {}
    for idx, (year, source, _row_source_required) in amount_columns.items():
        if source is None:
            amount = parse_money_to_rub(row_cells[idx].value)
            if amount != Decimal("0.00"):
                totals[year] = amount
    return totals
